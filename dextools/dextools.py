import cloudscraper
import re
import openpyxl

scraper = cloudscraper.create_scraper()
max_retry = 3
file_name = "sample.xlsx"  # input("Enter file name:")
time = input("Enter time (1s/5s/15s/1/3/5/15/30/60/120/...):")
res = time.upper()
q = "So11111111111111111111111111111111111111112"
# {api_point}/{pairAddress}?mc=1&bbn=300611376&res=1S&cb=2&q={chainId}
try:
    wb_obj = openpyxl.load_workbook(file_name)
    sheet_obj = wb_obj.active
    num_rows = sheet_obj.max_row - 1
except Exception as e:
    print(e)
    input("")


def getData(token_address, row):
    pairAddress = getPairAddress(token_address)
    if pairAddress == None:
        return None
    url = f"https://io.dexscreener.com/dex/chart/amm/v3/solamm/bars/solana/{pairAddress}?mc=1&res={res}&cb=1000&q={q}"

    response = scraper.request("GET", url)
    if response.status_code == 200:
        pattern = r"\d+\.\d{4,}"
        matches = re.findall(pattern, response.text)
        cleaned = matches[1::8]
        cleaned.append(matches[-1])
        cleaned.pop(0)
        to_add = cleaned
        for match in range(len(to_add)):
            sheet_obj.cell(row, match + 3).value = to_add[match]
        print(f"Inserted {len(cleaned)} value(s) for token {token_address}")


row = 2
column = 2
for x in range(num_rows):
    token_address = sheet_obj.cell(row, column).value
    getData(token_address, row)
    row += 1
try:
    wb_obj.save(file_name)
    wb_obj.close()
except Exception as e:
    print(f"Error while saving file {e}")
    input("")
input("Press enter to exit")
