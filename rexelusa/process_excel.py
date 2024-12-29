import openpyxl

file_name = "sample.xlsx"


def openExcel(file_name):
    try:
        wb_obj = openpyxl.load_workbook(file_name)
    except Exception as e:
        print(e, "creating it.")
        wb_obj = openpyxl.Workbook()
        sheet_obj = wb_obj.active
        for x in range(len(headers)):
            sheet_obj.cell(1, x + 1).value = headers[x]
    sheet_obj = wb_obj.active
    return wb_obj, sheet_obj


def close(wb_obj, file_name):
    wb_obj.save(file_name)
    wb_obj.close()


def addRow(sheet_obj, row):
    sheet_obj.insert_rows(row)


def writeCell(sheet_obj, row, column, data):
    sheet_obj.cell(row, column).value = data


def writeRow(sheet_obj, row, data):
    for x in range(len(data)):
        sheet_obj.cell(row, x + 1).value = data[x]


headers = [
    "productId",
    "url",
    "Website Source",
    "Manufacturer Name",
    "Manufacturer Part Number",
    "Manufacturer Description",
    "Supplier Name",
    "Supplier Part Number",
    "image",
    "Description",
    "Category",
    "Sub-category",
    "UNSPSC",
    "UPC",
    "Notes",
    "Price",
    "In Stock",
    "Substitute 1 MFG name",
    "Substitute 1 description",
    "Substitute 1 part number",
    "Substitute 2 MFG name",
    "Substitute 2 description",
    "Substitute 2 part number",
    "Substitute 3 MFG name",
    "Substitute 3 description",
    "Substitute 3 part number",
    "Substitute 4 MFG name",
    "Substitute 4 description",
    "Substitute 4 part number",
    "Attribute label 1",
    "Attribute value 1",
    "Attribute label 3",
    "Attribute value 3",
    "Attribute label 4",
    "Attribute value 4",
    "Attribute label 5",
    "Attribute value 5",
    "Attribute label 6",
    "Attribute value 6",
    "Attribute label 7",
    "Attribute value 7",
    "Attribute label 8",
    "Attribute value 8",
    "Attribute label 9",
    "Attribute value 9",
    "Attribute label 10",
    "Attribute value 10",
    "Attribute label 11",
    "Attribute value 11",
    "Attribute label 12",
    "Attribute value 12",
    "Attribute label 13",
    "Attribute value 13",
    "Attribute label 14",
    "Attribute value 14",
    "Attribute label 15",
    "Attribute value 15",
    "Attribute label 16",
    "Attribute value 16",
    "Attribute label 17",
    "Attribute value 17",
    "Attribute label 18",
    "Attribute value 18",
    "Attribute label 19",
    "Attribute value 19",
    "Attribute label 20",
    "Attribute value 20",
    "Attribute label 21",
    "Attribute value 21",
]
