import openpyxl

file_name = "rexel_Boxes & Enclosures.xlsx"


def openExcel(file_name):
    try:
        wb_obj = openpyxl.load_workbook(file_name, read_only=True)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        print(max_row)
        wb_obj.close
    except Exception as e:
        print(e)


openExcel(file_name)
