import openpyxl
import openpyxl.styles
from datetime import datetime


def openExcel(file_name):
    try:
        wb_obj = openpyxl.load_workbook(file_name)
    except Exception as e:
        print(e, "creating it.")
        wb_obj = openpyxl.Workbook()
        sheet_obj = wb_obj.active
        for x in range(len(headers)):
            sheet_obj.cell(1, x + 1).value = headers[x]
    sheet_obj = wb_obj.active
    return wb_obj, sheet_obj


def save(wb_obj, file_name):
    wb_obj.save(file_name)


def close(wb_obj, file_name):
    wb_obj.save(file_name)
    wb_obj.close()


def addRow(sheet_obj, row):
    sheet_obj.insert_rows(row)


def writeCell(sheet_obj, row, column, data):
    sheet_obj.cell(row, column).value = data


def writeRow(sheet_obj, row, data):
    for x in range(len(data)):
        sheet_obj.cell(row, x + 1).value = data[x]


def convertToNumber(sheet_obj):
    for cell in sheet_obj[sheet_obj.max_row]:
        if isinstance(cell.value, (float, int)):
            cell.number_format = openpyxl.styles.numbers.FORMAT_NUMBER


def merge_excel_files(file_list, output_file):
    # Create a new workbook and select the active worksheet
    merged_wb = openpyxl.Workbook()
    merged_ws = merged_wb.active

    for file in file_list:
        # Load the current workbook and select the active worksheet
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 and file != file_list[0]:
                # Skip the header row for all files except the first one
                continue

            # Append the row to the merged worksheet
            merged_ws.append(row)

    # Save the merged workbook
    merged_wb.save(output_file)

    # Get the current row count in the merged worksheet
    row_count = merged_ws.max_row
    print(f"Total row count: {row_count}")


from collections import defaultdict


def countRowsByScore(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheetname]

    # Initialize a dictionary to store counts of each score combination
    score_counts = defaultdict(int)

    # Iterate through rows, skipping the header
    print(ws.cell(2, 4).value, type(ws.cell(2, 3).value))
    for row in ws.iter_rows(min_row=2, values_only=True):
        final_score_t1, final_score_t2 = row[28], row[29]
        score_counts[(final_score_t1, final_score_t2)] += 1

    return score_counts


def convert_to_24_hour_format(time_str):
    if time_str == None:
        time_str = "12:00am"
    return datetime.strptime(time_str, "%I:%M%p").time()


def sort_excel_by_date(filename, sheetname, output_filename):
    # Dictionary to map month abbreviations to numbers
    month_to_number = {
        "Jan": 1,
        "Feb": 2,
        "Mar": 3,
        "Apr": 4,
        "May": 5,
        "Jun": 6,
        "Jul": 7,
        "Aug": 8,
        "Sep": 9,
        "Oct": 10,
        "Nov": 11,
        "Dec": 12,
    }
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheetname]
    year_col = 14
    month_col = 12
    day_col = 13
    hour_col = 15
    # Read header row
    headers = [cell.value for cell in ws[1]]

    # Read data rows and convert date columns to datetime objects
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        year = int(row[year_col - 1])  # Adjust index for zero-based indexing
        month = month_to_number[row[month_col - 1]]
        day = int(row[day_col - 1])
        hour_str = row[hour_col - 1]
        time = convert_to_24_hour_format(hour_str)
        date = datetime(year, month, day, time.hour, time.minute)
        data_rows.append((date, row))

    # Sort data rows by the date
    data_rows.sort(key=lambda x: x[0])

    # Create a new workbook and sheet for the sorted data
    new_wb = openpyxl.load_workbook(filename)
    new_ws = new_wb.active
    new_ws.title = "Sorted " + sheetname

    # Write header row to the new sheet
    for col_num, header in enumerate(headers, 1):
        new_ws.cell(row=1, column=col_num, value=header)

    # Write sorted data rows to the new sheet
    for row_num, (date, row) in enumerate(data_rows, 2):
        for col_num, value in enumerate(row, 1):
            new_ws.cell(row=row_num, column=col_num, value=value)

    # Save the new workbook
    new_wb.save(output_filename)
    new_wb.close()


from openpyxl import load_workbook


from openpyxl import load_workbook, Workbook


def remove_duplicates(input_file, output_file, url_column_name="URL"):
    # Load the workbook and the first sheet
    workbook = load_workbook(input_file)
    sheet = workbook.active

    # Get the header row
    header = [cell.value for cell in sheet[1]]

    # Find the column index of the URL column
    if url_column_name not in header:
        raise ValueError(f"Column '{url_column_name}' not found in header")

    url_column_index = header.index(url_column_name) + 1

    # Use a dictionary to track unique URLs
    seen_urls = {}
    deduplicated_rows = [header]  # Start with the header row

    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[url_column_index - 1]
        if url not in seen_urls:
            seen_urls[url] = True
            deduplicated_rows.append(row)

    # Create a new workbook for deduplicated data
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # Write deduplicated data to the new sheet
    for row in deduplicated_rows:
        new_sheet.append(row)

    # Save the new workbook
    new_workbook.save(output_file)
    print(f"Duplicates removed. Saved to {output_file}")


headers = [
    "Team 1",
    "Team 2",
    "Record T1 win",
    "Record T1 loss",
    "Record T1 tie",
    "Record T2 win",
    "Record T2 loss",
    "Record T2 tie",
    "Coach team 1",
    "Coach team 2",
    "Day name",
    "Month",
    "Day",
    "Year",
    "Start time",
    "Stadium",
    "Quarter 1 T1",
    "Quarter 1 T2",
    "Quarter 2 T1",
    "Quarter 2 T2",
    "Quarter 3 T1",
    "Quarter 3 T2",
    "Quarter 4 T1",
    "Quarter 4 T2",
    "OT 1 T1",
    "OT 1 T2",
    "OT 2 T1",
    "OT 2 T2",
    "Final T1",
    "Final T2",
    "Weather temp",
    "Weather humidity ",
    "Weather wind",
    "Vegas line team",
    "Vegas line number",
    "Over/Under",
    "Over/Under number",
    "advanced passing team 1 player",
    "cmp",
    "att",
    "yds",
    "sk",
    "advanced passing team 2 player",
    "cmp",
    "att",
    "yds",
    "sk",
    "advanced rushing team 1 player",
    "att",
    "yds",
    "td",
    "advanced rushing team 2 player",
    "att",
    "yds",
    "td",
    "advanced receiving team 1 player",
    "tgt",
    "rec",
    "yds",
    "td",
    "advanced receiving team 2 player",
    "tgt",
    "rec",
    "yds",
    "td",
    "URL",
]

if __name__ == "__main__":
    # Merge files
    file_list = ["nfl3.xlsx", "nfl45_60.xlsx", "nfl60_100.xlsx", "nfl100.xlsx"]
    output_file = "nfl_full.xlsx"
    sheetname = "Sheet"
    # merge_excel_files(file_list, output_file)

    # Remove duplicate based on column name "URL"
    filename = "nfl_full_nodeplucate.xlsx"
    # remove_duplicates(output_file, filename)

    # Sort by date
    output_filename = "nfl_full_sorted.xlsx"
    sort_excel_by_date(filename, sheetname, output_filename)
