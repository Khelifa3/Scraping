import cloudscraper
from bs4 import BeautifulSoup
import re
import process_excel
import openpyxl
import time

email_pattern = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
scraper = cloudscraper.create_scraper()
# Main scraping
file_name = "./register1.xlsx"
wb_obj, sheet_obj = process_excel.openExcel(file_name)
sheet_obj.title = "Charities"
rel_obj, rel_sheet = process_excel.openExcel("religionCharities.xlsx")
rel_sheet.title = "Religion Charities"

process_excel.close(wb_obj, file_name)


def getRegister():
    try:
        source_obj = openpyxl.load_workbook("./community/register1.xlsx")
    except Exception as e:
        print(e)
    source_sheet = source_obj.active
    max_rows = 14150
    for row in range(1800, max_rows):
        name = source_sheet.cell(row, 2).value
        domaine = source_sheet.cell(row, 5).value  # religion
        if "religion" not in domaine.lower():
            continue
        address = source_sheet.cell(row, 6).value
        email = getEmail(name)
        if email == "":
            continue
        entity = [name, email, address]

        new_row = rel_sheet.max_row + 1
        process_excel.writeRow(rel_sheet, new_row, entity)
        process_excel.save(rel_obj, "religionCharities.xlsx")
        print(f"Saved row {row} in religion")


def getEmail(name):
    name = "+".join(name.split(" "))
    url = f"https://www.google.com/search?q=email+{name}"
    response = scraper.get(url)
    print(response.status_code)
    if response.status_code == 429:
        time.sleep(60)
    soup = BeautifulSoup(response.text, "html.parser")
    try:
        email = re.findall(email_pattern, soup.text.strip())[0]
    except:
        email = ""
    print(email)
    time.sleep(2)
    return email


getRegister()
