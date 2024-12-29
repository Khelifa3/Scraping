import openpyxl
import os
import pandas as pd

file_name = "sample.xlsx"

headers = ["Name", "Email", "Address"]


def openExcel(file_name):
    try:
        wb_obj = openpyxl.load_workbook(file_name)
    except Exception as e:
        print(e, "creating it.")
        wb_obj = openpyxl.Workbook()
        sheet_obj = wb_obj.active
        for x in range(len(headers)):
            sheet_obj.cell(1, x + 1).value = headers[x]
    sheet_obj = wb_obj.active
    return wb_obj, sheet_obj


def close(wb_obj, file_name):
    wb_obj.save(file_name)
    wb_obj.close()


def save(wb_obj, file_name):
    wb_obj.save(file_name)


def addRow(sheet_obj, row):
    sheet_obj.insert_rows(row)


def writeCell(sheet_obj, row, column, data):
    sheet_obj.cell(row, column).value = data


def writeRow(sheet_obj, row, data):
    for x in range(len(data)):
        sheet_obj.cell(row, x + 1).value = data[x]


def count_rows_in_xlsx_files(folder_path):
    total_rows = 0

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".xlsx"):
                file_path = os.path.join(root, file)
                try:
                    df = pd.read_excel(file_path)
                    num_rows = df.shape[0]  # Number of rows
                    print(file_path, num_rows)
                    total_rows += num_rows
                except Exception as e:
                    print(f"Error reading {file_path}: {e}")

    return total_rows


import os
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook


def merge_xlsx_files(folder_path, output_file, sheet_name):
    merged_data = []

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".xlsx"):
                file_path = os.path.join(root, file)
                try:
                    wb = load_workbook(file_path)
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        merged_data.append(row)
                except Exception as e:
                    print(f"Error reading {file_path}: {e}")

    # Remove duplicate and empty rows
    merged_data = list(filter(lambda x: any(x), merged_data))  # Remove empty rows
    merged_data = list(
        dict.fromkeys(merged_data)
    )  # Remove duplicates while preserving order

    # Write merged data to output file
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name
    for row in merged_data:
        new_ws.append(row)

    new_wb.save(output_file)
    print(f"All files have been merged into {output_file}")


# folder_path = 'your/folder/path/here'  # Replace with the path to your folder
# output_file = 'merged_output.xlsx'     # Replace with your desired output file name
# merge_xlsx_files(folder_path, output_file)


def count_rows_in_xlsx_files(folder_path):
    total_rows = 0

    for file in os.listdir(folder_path):
        if file.endswith(".xlsx"):
            file_path = os.path.join(folder_path, file)
            try:
                workbook = load_workbook(file_path, data_only=True)
                for sheet in workbook.worksheets:
                    total_rows += sheet.max_row
                    print(file_path, sheet.max_row)
            except Exception as e:
                print(f"Error reading {file_path}: {e}")

    return total_rows


# folder_path = "."  # Replace with the path to your folder
# total_rows = count_rows_in_xlsx_files(folder_path)
# print(f"Total number of rows in all .xlsx files: {total_rows}")
def copyFile(src, dest):

    # Load the source and destination workbooks
    source_wb = openpyxl.load_workbook("source.xlsx")
    dest_wb = openpyxl.load_workbook("destination.xlsx")

    # Select the active sheets (or specify by name)
    source_sheet = source_wb.active
    dest_sheet = dest_wb.active

    # Specify the columns you want to copy (e.g., columns A and B)
    columns_to_copy = ["A", "B"]

    # Loop through each row and copy the specified columns
    for row in range(1, source_sheet.max_row + 1):
        for col in columns_to_copy:
            dest_sheet[f"{col}{row}"] = source_sheet[f"{col}{row}"].value

    # Save the destination workbook
    dest_wb.save("destination.xlsx")

    print("Columns copied successfully!")


def merge(a, b, new):
    # Load the first workbook
    wb1 = openpyxl.load_workbook(a + ".xlsx")
    ws1 = wb1.active

    # Load the second workbook
    wb2 = openpyxl.load_workbook(b + ".xlsx")
    ws2 = wb2.active

    # Copy the data from ws2 to ws1
    for row in ws2.iter_rows(values_only=True):
        ws1.append(row)

    # Save the merged workbook to a new file
    wb1.save(new + ".xlsx")


def finalMerge():

    # List of Excel files to be merged
    file_list = [
        "Arts.xlsx",
        "Childcare.xlsx",
        "communitySheet.xlsx",
        "Education.xlsx",
        "Healthcare.xlsx",
        "Legal.xlsx",
        "religionSheet.xlsx",
        "Sports.xlsx",
        "Tourism.xlsx",
        "Transport and hostpitality.xlsx",
    ]

    # Create a new workbook
    merged_wb = openpyxl.Workbook()
    merged_wb.remove(merged_wb.active)  # Remove the default sheet

    for file_name in file_list:
        # Load each workbook
        wb = openpyxl.load_workbook(file_name)
        for sheet_name in wb.sheetnames:
            # Get the sheet from the workbook
            sheet = wb[sheet_name]

            # Create a new sheet in the merged workbook with the same title
            new_sheet = merged_wb.create_sheet(title=sheet_name)

            # Copy the data from the sheet to the new_sheet
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)

    # Save the merged workbook to a new file
    merged_wb.save("irelandData.xlsx")


def finalCount(f_name):
    # Load the workbook
    wb = openpyxl.load_workbook(f_name)

    # Dictionary to store sheet name and row count
    row_counts = {}
    total = 0
    # Iterate through each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        row_counts[sheet_name] = sheet.max_row
        total += sheet.max_row

    # Print the row counts
    for sheet_name, row_count in row_counts.items():
        print(f'Sheet "{sheet_name}" has {row_count} rows')
    print(f"Total: {total}")


def clean_excel_sheets(input_file: str, output_file: str):
    # Load the Excel file
    excel_data = pd.ExcelFile(input_file)
    cleaned_data = {}

    for sheet_name in excel_data.sheet_names:
        # Read each sheet into a DataFrame
        df = excel_data.parse(sheet_name)

        # Drop rows with any NaN values and remove duplicates
        df_cleaned = df.dropna().drop_duplicates()
        # Store the cleaned DataFrame in a dictionary
        cleaned_data[sheet_name] = df_cleaned

    # Write cleaned data to a new Excel file
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, df_cleaned in cleaned_data.items():
            df_cleaned.to_excel(writer, sheet_name=sheet_name, index=False)


if __name__ == "__main__":

    # folder_path = "."  # Replace with the path to your folder
    # total_rows = count_rows_in_xlsx_files(folder_path)
    # print(f"Total number of rows in all .xlsx files: {total_rows}")
    sheet_names = [
        "Education",  # 4200
        "Childcare and Early Years Service",  # 5000
        "Healthcare and Social Care",  # 750
        "Youth Work and Sports",  # 5250
        "Religious and Faith-Based Organ",  # 1500
        "Community and Voluntary Sector",  # 4500
        "Transport and Hospitality",  # 340
        "Performing Arts and Entertainment",  # 450
        "Tourism and Exchange Programs",  # 110
        "Legal and Advocacy Services",  # 200
    ]
    # merge_xlsx_files("./legal", "Legal.xlsx", "Legal and Advocacy Services")
    merge("Childcare", "childcareonline", "Childcare")
    input_file = "irelandData.xlsx"  # Replace with your input file path
    output_file = "output_file.xlsx"  # Replace with your desired output file path
    finalMerge()
    # finalCount(input_file)
    clean_excel_sheets(input_file, output_file)
    finalCount("output_file.xlsx")
