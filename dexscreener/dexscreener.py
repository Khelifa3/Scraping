import cloudscraper
import re
import openpyxl

scraper = cloudscraper.create_scraper()
max_retry = 3
file_name = "DATOS HISTORICOS MIGRACION MEMECOINS (1).xlsx"  # input("Enter file name:")
time = input("Enter time (1s/5s/15s/1/3/5/15/30/60/120/...):")
res = time.upper()
q = "So11111111111111111111111111111111111111112"
# {api_point}/{pairAddress}?mc=1&bbn=300611376&res=1S&cb=2&q={chainId}
try:
    wb_obj = openpyxl.load_workbook(file_name)
    sheet_obj = wb_obj.active
    num_rows = sheet_obj.max_row - 1
except Exception as e:
    print(e)
    input("")


def getPairAddress(token_address, retry=0):
    try:
        response = scraper.get(
            f"https://api.dexscreener.com/latest/dex/tokens/{token_address}"
        )
        if response.status_code != 200:
            if retry < max_retry:
                print(
                    f"Error getting {token_address} {response.status_code} retrying..."
                )
                retry += 1
                getPairAddress(token_address, retry)
            else:
                print(f"Failed to get {token_address} {response.status_code}")
                return None
        data = response.json()
        if data["pairs"] == None:
            print(f"No pair found for {token_address}")
            return None
        else:
            pairAddress = data["pairs"][0]["pairAddress"]
            return pairAddress
    except Exception as e:
        if retry < max_retry:
            print(f"Error getting {token_address} {e} retrying...")
            retry += 1
            getPairAddress(token_address, retry)
        else:
            print(f"Failed to get {token_address} {e}")
            return None


def getData(token_address, row):
    pairAddress = getPairAddress(token_address)
    if pairAddress == None:
        return None
    url = f"https://io.dexscreener.com/dex/chart/amm/v3/solamm/bars/solana/{pairAddress}?mc=1&res={res}&cb=1000&q={q}"

    response = scraper.request("GET", url)
    if response.status_code == 200:
        pattern = r"\d+\.\d{4,}"
        matches = re.findall(pattern, response.text)
        cleaned = matches[1::8]
        cleaned.append(matches[-1])
        cleaned.pop(0)
        to_add = cleaned
        for match in range(len(to_add)):
            sheet_obj.cell(row, match + 3).value = to_add[match]
        print(f"Inserted {len(cleaned)} value(s) for token {token_address}")


row = 2
column = 2
for x in range(num_rows):
    token_address = sheet_obj.cell(row, column).value
    getData(token_address, row)
    row += 1
try:
    wb_obj.save(file_name)
except Exception as e:
    print(f"Error while saving file {e}")
    input("")
input("Press enter to exit")
